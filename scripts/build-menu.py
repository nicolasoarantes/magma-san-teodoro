#!/usr/bin/env python3
"""Regenerate src/data/menu.ts and public/menu/magma-menu.pdf from cardapio.xlsx.

Usage:  python scripts/build-menu.py
Deps:   pip install openpyxl reportlab
"""
import os, re, json
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "cardapio.xlsx")
TS_OUT = os.path.join(ROOT, "src", "data", "menu.ts")
PDF_OUT = os.path.join(ROOT, "public", "menu", "magma-menu.pdf")
QR_OUT = os.path.join(ROOT, "public", "menu", "qr-menu.png")
SITE_MENU_URL = "https://magma-san-teodoro.vercel.app/menu"  # TODO: ajuste para o domínio real


def isnum(x):
    try:
        if x is None or isinstance(x, bool):
            return False
        float(x)
        return True
    except Exception:
        return False


def normalize():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))[1:]
    items = []
    for r in rows:
        if not r or not r[0]:
            continue
        cat, sub, name = r[0], r[1], r[2]
        rest = list(r[3:7]) + [None] * (4 - len(r[3:7]))
        price, pidx = None, None
        for i, c in enumerate(rest):
            if isnum(c):
                price, pidx = float(c), i
                break
        unit = "€"
        if pidx is not None and pidx + 1 < len(rest) and isinstance(rest[pidx + 1], str) and rest[pidx + 1].strip():
            unit = rest[pidx + 1].strip()
        pre = [c for j, c in enumerate(rest) if (pidx is None or j < pidx) and isinstance(c, str) and c.strip()]
        details, desc = [], []
        for c in pre:
            (details if re.match(r"^\s*\d+\s*[lL]\s*$", c) or len(c) <= 4 else desc).append(c.strip())
        items.append({
            "category": cat, "subcategory": sub, "name": str(name).strip(),
            "description": " ".join(desc), "details": " ".join(details),
            "price": price, "unit": unit,
        })
    return items


def esc(s):
    return (s or "").replace("\\", "\\\\").replace('"', '\\"')


def write_ts(items):
    order = []
    for it in items:
        if it["category"] not in order:
            order.append(it["category"])
    L = ["// AUTO-GENERATED from cardapio.xlsx — do not edit by hand.",
         "// To regenerate: python scripts/build-menu.py",
         "export interface MenuItem {",
         "  category: string;", "  subcategory: string | null;", "  name: string;",
         "  description: string;", "  details: string;", "  price: number | null;",
         "  unit: string;", "}\n",
         "export const menuCategories: string[] = ["]
    L += [f'  "{esc(c)}",' for c in order]
    L += ["];\n", "export const menuItems: MenuItem[] = ["]
    for it in items:
        sub = f'"{esc(it["subcategory"])}"' if it["subcategory"] else "null"
        p = "null" if it["price"] is None else (str(int(it["price"])) if float(it["price"]).is_integer() else str(it["price"]))
        L.append(f'  {{ category: "{esc(it["category"])}", subcategory: {sub}, '
                 f'name: "{esc(it["name"])}", description: "{esc(it["description"])}", '
                 f'details: "{esc(it["details"])}", price: {p}, unit: "{esc(it["unit"])}" }},')
    L.append("];")
    open(TS_OUT, "w", encoding="utf-8").write("\n".join(L))
    print("wrote", TS_OUT, f"({len(items)} items)")


def write_pdf(items):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
    from reportlab.graphics.barcode import qr
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics import renderPM

    q = qr.QrCodeWidget(SITE_MENU_URL)
    b = q.getBounds(); wq, hq = b[2] - b[0], b[3] - b[1]; sz = 600
    d = Drawing(sz, sz, transform=[sz / wq, 0, 0, sz / hq, 0, 0]); d.add(q)
    renderPM.drawToFile(d, QR_OUT, "PNG")
    print("wrote", QR_OUT)

    order, data = [], {}
    for it in items:
        if it["category"] not in data:
            data[it["category"]] = {}
            order.append(it["category"])
        data[it["category"]].setdefault(it["subcategory"] or "_", []).append(it)

    BLACK = colors.HexColor("#0a0807"); ORANGE = colors.HexColor("#FF6A00")
    GOLD = colors.HexColor("#F5A623"); CREAM = colors.HexColor("#FFF4E2"); GREY = colors.HexColor("#b8b0a6")
    c = canvas.Canvas(PDF_OUT, pagesize=A4); W, H = A4; y = [0]

    def bg():
        c.setFillColor(BLACK); c.rect(0, 0, W, H, fill=1, stroke=0)

    def newpage():
        bg(); y[0] = H - 22 * mm
        c.setStrokeColor(ORANGE); c.setLineWidth(1.2); c.line(18 * mm, 18 * mm, W - 18 * mm, 18 * mm)
        c.setFont("Helvetica", 7); c.setFillColor(GREY)
        c.drawCentredString(W / 2, 12 * mm, "MAGMA — Via San Francesco d'Assisi 15, San Teodoro (SS) — +39 340 139 6582")

    def need(n):
        if y[0] - n < 24 * mm:
            c.showPage(); newpage()

    def fmt(it):
        if it["price"] is None:
            return ""
        p = int(it["price"]) if float(it["price"]).is_integer() else it["price"]
        ex = (it["unit"] or "€").replace("€", "").strip()
        return f"€ {p}" + (f" {ex}" if ex else "")

    bg()
    c.setFillColor(ORANGE); c.setFont("Helvetica-Bold", 72); c.drawCentredString(W / 2, H - 120 * mm, "MAGMA")
    c.setFillColor(CREAM); c.setFont("Helvetica", 13); c.drawCentredString(W / 2, H - 132 * mm, "CHURRASCARIA  •  STEAKHOUSE  •  AMERICAN BAR")
    c.setFillColor(GOLD); c.setFont("Helvetica-Oblique", 12); c.drawCentredString(W / 2, H - 148 * mm, "Una notte, tutto da vivere.")
    c.showPage(); newpage()
    for cat in order:
        need(30 * mm)
        c.setFillColor(ORANGE); c.setFont("Helvetica-Bold", 20); c.drawString(18 * mm, y[0], cat.upper())
        y[0] -= 3 * mm; c.setStrokeColor(ORANGE); c.setLineWidth(0.8); c.line(18 * mm, y[0], W - 18 * mm, y[0]); y[0] -= 8 * mm
        for sub, lst in data[cat].items():
            if sub != "_":
                need(12 * mm); c.setFillColor(GOLD); c.setFont("Helvetica-Bold", 11); c.drawString(20 * mm, y[0], sub.upper()); y[0] -= 6 * mm
            for it in lst:
                need(14 * mm)
                nm = it["name"] + (f"  ·  {it['details']}" if it["details"] else "")
                c.setFillColor(CREAM); c.setFont("Helvetica-Bold", 11); c.drawString(20 * mm, y[0], nm)
                c.setFillColor(ORANGE); c.setFont("Helvetica-Bold", 11); c.drawRightString(W - 18 * mm, y[0], fmt(it)); y[0] -= 5 * mm
                if it["description"]:
                    c.setFillColor(GREY); c.setFont("Helvetica", 8.5); line = ""
                    for wd in it["description"].split():
                        if c.stringWidth(line + " " + wd, "Helvetica", 8.5) > (W - 40 * mm):
                            c.drawString(20 * mm, y[0], line); y[0] -= 4 * mm; need(8 * mm); line = wd
                        else:
                            line = (line + " " + wd).strip()
                    if line:
                        c.drawString(20 * mm, y[0], line); y[0] -= 4 * mm
                y[0] -= 2.5 * mm
        y[0] -= 4 * mm
    c.showPage(); c.save()
    print("wrote", PDF_OUT)


if __name__ == "__main__":
    items = normalize()
    write_ts(items)
    try:
        write_pdf(items)
    except Exception as e:  # PDF is optional; menu.ts is the critical output
        print("PDF/QR skipped:", e)
